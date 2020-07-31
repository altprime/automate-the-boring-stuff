'''
blank row inserter

take two integers as args
N -> starting row
M -> number of rows to insert
starting at N, insert M blank rows into spreadsheet

'''

import openpyxl,sys
from openpyxl.utils import column_index_from_string 

if len(sys.argv)!=4:
    print('Error!')
    sys.exit()
N = int(sys.argv[1])
M = int(sys.argv[2])

file = sys.argv[3]

wb = openpyxl.load_workbook(file)
wbNew = openpyxl.Workbook()

sheet1 = wb.active
sheet2 = wbNew['Sheet']

for rowObj in sheet1.rows:
    for cellObj in rowObj:
        #print(cellObj.column)
       
        if cellObj.row<N:
            sheet2.cell(row=cellObj.row,column=int(cellObj.column)).value = cellObj.value
        else:
            sheet2.cell(row=cellObj.row+M,column=int(cellObj.column)).value = cellObj.value

wbNew.save('exampleInseter.xlsx')
print('Done!')