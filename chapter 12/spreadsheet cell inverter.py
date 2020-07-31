'''
spreadsheet cewll inverter

in short... transpose it all

'''

import openpyxl

wb = openpyxl.load_workbook('invertThis.xlsx')
sheet = wb['Sheet1']

wbNew = openpyxl.Workbook()
sheet2 = wbNew['Sheet']


for rowObj in sheet.rows:
    for cellObj in rowObj:
        sheet2.cell(row=cellObj.row,column=cellObj.column).value = sheet.cell(row=cellObj.column,column=cellObj.row).value


wbNew.save('invertedSheet.xlsx')
