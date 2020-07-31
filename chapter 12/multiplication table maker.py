'''
multiplication table maker

take number N from cmd line and make an NxN matrix/table 


	1	2	3	4	5
1	1	2	3	4	5
2	2	4	6	8	10
3	3	6	9	12	15
4	4	8	12	16	20
5	5	10	15	20	25


'''


import openpyxl, sys
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

fontObj = Font(name='Georgia',bold=True)
n = int(sys.argv[1])
for i in range(1, n+1):
    sheet['A' + str(i+1)] = i
    sheet.cell(row=1,column=i+1).value = i
    
    sheet['A'+ str(i+1)].font = fontObj
    sheet.cell(row=1,column=i+1).font = fontObj

for j in range(1, n+1):
    for k in range(1, n+1):
        sheet.cell(row = j+1 ,column= k+1).value = j*k   


wb.save('multiplication tabel.xlsx')
 
