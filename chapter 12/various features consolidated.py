'''
consolidated codes

'''
# setting font
from openpyxl.styles import Font, Style
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
italic24Font = Font(size=24, italic=True)  
styleObj = Style(font=italic24Font) 
sheet['A1'].style = styleObj 
sheet['A1'] = 'Hello world!' 
wb.save('styled.xlsx')

# formulas
import openpyxl
wb = openpyxl.Workbook() 
sheet = wb.get_active_sheet() 
sheet['A1'] = 200 
sheet['A2'] = 300 
sheet['A3'] = '=SUM(A1:A2)' 
wb.save('writeFormula.xlsx')

import openpyxl 
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx') 
sheet = wbFormulas.get_active_sheet() 
sheet['A3'].value # =SUM(A1:A2)
wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.get_active_sheet() >>> sheet['A3'].value # 500

# adjusting rows and columns
import openpyxl 
wb = openpyxl.Workbook() 
sheet = wb.get_active_sheet() 
sheet['A1'] = 'Tall row' 
sheet['B2'] = 'Wide column' 
sheet.row_dimensions[1].height = 70 
sheet.column_dimensions['B'].width = 20 
wb.save('dimensions.xlsx')

# merging and unmergins cells
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet() 
sheet.merge_cells('A1:D3') 
sheet['A1'] = 'Twelve cells merged together.' 
sheet.merge_cells('C5:D5') 
sheet['C5'] = 'Two merged cells.' 
wb.save('merged.xlsx')


wb = openpyxl.load_workbook('merged.xlsx') 
sheet = wb.get_active_sheet() 
sheet.unmerge_cells('A1:D3') 
sheet.unmerge_cells('C5:D5') 
wb.save('merged.xlsx')

# charts
import openpyxl 
wb = openpyxl.Workbook() 
sheet = wb.get_active_sheet() 
for i in range(1, 11):
    sheet['A' + str(i)] = i
    
refObj = openpyxl.charts.Reference(sheet, (1, 1), (10, 1))

seriesObj = openpyxl.charts.Series(refObj, title='First series')

chartObj = openpyxl.charts.BarChart() 
chartObj.append(seriesObj)

chartObj.drawing.top = 50
chartObj.drawing.left = 100 
chartObj.drawing.width = 300 
chartObj.drawing.height = 200

sheet.add_chart(chartObj)
wb.save('sample chart.xlsx')
